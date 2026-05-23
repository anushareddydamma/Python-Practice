import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2,sheet.max_row + 1):
        cell1 = sheet.cell(row,3)
        cell2 = sheet.cell(row,4)
        total_price = cell1.value * cell2.value
        sheet.cell(row, 5).value = total_price

    values = Reference(sheet,
                  min_row = 2,
                  max_row = sheet.max_row,
                  min_col = 3,
                  max_col = 3,)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'F2')

    wb.save('filename')
